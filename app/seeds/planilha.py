"""Migração ÚNICA da planilha do brigada (carga inicial — não é a feature de lote).

Lê `1aESCALAS SERVIÇO_QG _2025.xlsx` (a escala viva) e cria: OMs, militares
(só nome+posto+OM; identidade/antiguidade ficam p/ a ficha), as 5 escalas
concorrentes, 1 posto por escala, as participações e o ESTADO DA FILA (um serviço
sintético na última data PRETA/VERMELHA de cada militar, p/ o motor continuar de
onde a planilha parou).

    python -m app.seeds.planilha            # DRY-RUN (não grava; mostra o resumo)
    python -m app.seeds.planilha --commit   # grava de fato

Reconciliação decidida com o gestor (2026-07-24):
  - postos: normaliza maiúsculas + sufixo feminino (– F/-F) + alias TC/TEN CEL;
    '4º Sgt' é digitação -> '3º Sgt'; 'GRAD' é lixo de cabeçalho -> ignora a linha.
  - OMs: funde variantes (1º C Geo=1º CGeo; B Adm Ap/3ª RM=B Adm Ap/3;
    Cmdo CMS – E5=Cmdo CMS; maiúsculas/‘/’); 'Coluna1' é lixo -> ignora.
  - militar único por (NOME DE GUERRA + OM).
  - as 5 escalas entram como mutuamente concorrentes (ajustável depois na gestão).
"""
from __future__ import annotations

import argparse
import re
import sys
from collections import defaultdict
from datetime import date, datetime, time, timedelta
from pathlib import Path

import openpyxl
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.database import SessionLocal
from app.domain.models import Cor
from app.models.escala import Escala, EscalaConcorrente, Participacao, Posto
from app.models.militar import Militar
from app.models.referencia import OrganizacaoMilitar, PostoGraduacao
from app.models.servico import Servico
from app.seeds.dados import POSTOS_GRADUACAO

ARQUIVO = "1aESCALAS SERVIÇO_QG _2025.xlsx"
SENTINELA = date(2001, 1, 1)   # "nunca serviu" na planilha

# aba -> (nome da escala, tem_preta, tem_vermelha)
ESCALAS = {
    "Supe_Dia_Gu": ("Supervisor de Dia da Guarda", True, True),
    "Of_Dia_QG": ("Oficial de Dia do QG", True, True),
    "Adj_Of_Dia": ("Adjunto do Oficial de Dia", True, True),
    "Perm_Port": ("Permanência do Portão", True, True),
    "Museu": ("Museu", False, True),     # só vermelha (regra 4.5/6)
}

# --- normalização de posto/graduação ---
_SIGLAS = {s.upper(): s for (s, *_rest) in POSTOS_GRADUACAO}
_ALIAS_POSTO = {"TC": "Ten Cel", "TEN CEL": "Ten Cel", "CAP": "Cap", "MAJ": "Maj",
                "4º SGT": "3º Sgt"}   # '4º Sgt' é digitação -> 3º Sgt


def norm_posto(raw) -> str | None:
    s = re.sub(r"\s*[–-]\s*F$", "", str(raw).strip())   # sufixo feminino
    s = re.sub(r"\s+", " ", s).strip()
    key = s.upper()
    if key in _SIGLAS:
        return _SIGLAS[key]
    if key in _ALIAS_POSTO:
        return _ALIAS_POSTO[key]
    return None    # 'GRAD' e afins caem aqui -> linha ignorada


# --- normalização/fusão de OM ---
def _chave_om(raw: str) -> str:
    s = re.sub(r"\s+", " ", str(raw).strip()).upper().replace("/", " ")
    return re.sub(r"\s+", " ", s).strip()


# chave normalizada -> sigla canônica de exibição (fusões confirmadas pelo gestor)
_OM_CANON = {
    "1º C GEO": "1º CGeo", "1º CGEO": "1º CGeo",
    "B ADM AP 3": "B Adm Ap/3", "B ADM AP 3ª RM": "B Adm Ap/3",
    "CMDO CMS": "Cmdo CMS", "CMDO CMS – E5": "Cmdo CMS", "CMDO CMS - E5": "Cmdo CMS",
    "CMDO 3ª RM": "Cmdo 3ª RM", "CMDO 6ª DE": "Cmdo 6ª DE",
    "CMDO 4º GPT E": "Cmdo 4º Gpt E", "CMDO 3º GPT LOG": "Cmdo 3º Gpt Log",
    "CPOR PA": "CPOR/PA", "CRO 3": "CRO/3", "3º CGCFEX": "3º CGCFEx",
    "3º B COM GE": "3º B Com GE", "3º GPT LOG": "3º Gpt Log",
    "8º B LOG": "8º B Log", "1º CTA": "1º CTA", "3º RCG": "3º RCG",
    "CMPA": "CMPA", "PMPA": "PMPA",
}
_OM_LIXO = {"COLUNA1"}


def canon_om(raw) -> str | None:
    k = _chave_om(raw)
    if k in _OM_LIXO:
        return None
    return _OM_CANON.get(k, str(raw).strip())   # desconhecida: mantém como veio


def _como_data(v) -> date | None:
    if isinstance(v, datetime):
        v = v.date()
    if isinstance(v, date) and v != SENTINELA:
        return v
    return None


# --- leitura da planilha ---
class Registro:
    __slots__ = ("nome", "posto", "om", "preta", "vermelha", "escala")

    def __init__(self, nome, posto, om, preta, vermelha, escala):
        self.nome, self.posto, self.om = nome, posto, om
        self.preta, self.vermelha, self.escala = preta, vermelha, escala


def ler_planilha(caminho: Path) -> tuple[list[Registro], list[str]]:
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    registros: list[Registro] = []
    avisos: list[str] = []
    for aba in ESCALAS:
        ws = wb[aba]
        header = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]

        def idx(*names):
            for i, h in enumerate(header):
                if h.upper() in names:
                    return i
            return None

        ip, inome, iom = idx("POSTO", "GRAD"), idx("NOME"), idx("OM")
        ipreta, iverm = idx("PRETA"), idx("VERMELHA", "VERMEL", "VERM")
        vazio = 0
        for row in ws.iter_rows(min_row=2, max_row=400, values_only=True):
            nome = row[inome] if inome is not None and inome < len(row) else None
            if not nome or not str(nome).strip():
                vazio += 1
                if vazio > 15:
                    break
                continue
            vazio = 0
            posto = norm_posto(row[ip]) if ip is not None and row[ip] else None
            om = canon_om(row[iom]) if iom is not None and row[iom] else None
            if posto is None or om is None:
                avisos.append(f"[{aba}] linha ignorada (posto/OM inválido): "
                              f"{str(nome).strip()!r} posto={row[ip]!r} om={row[iom]!r}")
                continue
            preta = _como_data(row[ipreta]) if ipreta is not None and ipreta < len(row) else None
            verm = _como_data(row[iverm]) if iverm is not None and iverm < len(row) else None
            registros.append(Registro(str(nome).strip(), posto, om, preta, verm, aba))
    return registros, avisos


# --- consolidação ---
def consolidar(registros: list[Registro]):
    """Agrupa por militar (nome de guerra + OM) e por escala."""
    oms = sorted({r.om for r in registros})
    militares: dict[tuple[str, str], dict] = {}
    avisos: list[str] = []
    # chave = nome de guerra + OM + posto: sem CPF/identidade, incluir o posto
    # evita FUNDIR pessoas distintas de mesmo nome/OM (ex.: um Maj e um 1º Ten
    # "JULIANA"). No pior caso gera um duplicado, que o brigada funde depois.
    for r in registros:
        chave = (r.nome.upper(), r.om, r.posto)
        if chave not in militares:
            militares[chave] = {"nome": r.nome, "om": r.om, "posto": r.posto}
    return oms, militares, avisos


def resumir(registros, oms, militares, avisos_ler, avisos_cons):
    por_escala = defaultdict(lambda: {"part": 0, "preta": 0, "verm": 0})
    postos_cont = defaultdict(int)
    for r in registros:
        e = por_escala[r.escala]
        e["part"] += 1
        if r.preta:
            e["preta"] += 1
        if r.vermelha:
            e["verm"] += 1
    for m in militares.values():
        postos_cont[m["posto"]] += 1

    print("=" * 64)
    print("DRY-RUN — migração da planilha (nada foi gravado)")
    print("=" * 64)
    print(f"\nMilitares únicos (nome de guerra + OM): {len(militares)}")
    print("  por posto:", dict(sorted(postos_cont.items(), key=lambda x: -x[1])))
    print(f"\nOMs a criar: {len(oms)}")
    print("  ", ", ".join(oms))
    n_esc = len(ESCALAS)
    print(f"\nEscalas a criar: {n_esc} (1 posto cada); concorrência: "
          f"{n_esc * (n_esc - 1) // 2} pares (todas concorrentes)")
    for aba, (nome, tp, tv) in ESCALAS.items():
        e = por_escala[aba]
        cor = "preta+vermelha" if tp and tv else "só vermelha"
        print(f"  - {nome:32} [{cor:14}] participações={e['part']:3}  "
              f"estado fila: preta={e['preta']:3} vermelha={e['verm']:3}")
    total_serv = sum(e["preta"] + e["verm"] for e in por_escala.values())
    print(f"\nServiços sintéticos (estado da fila) a gravar: ~{total_serv}")
    if avisos_ler or avisos_cons:
        print(f"\nAVISOS ({len(avisos_ler) + len(avisos_cons)}):")
        for a in avisos_ler + avisos_cons:
            print("  •", a)


# --- gravação ---
def gravar(db: Session, registros, oms, militares) -> dict:
    pg = {p.sigla: p.id for p in db.scalars(select(PostoGraduacao))}
    faltando = {m["posto"] for m in militares.values()} - set(pg)
    if faltando:
        raise SystemExit(f"posto/graduação não encontrado no banco (rode o seed): {faltando}")

    # OMs
    om_id = {o.sigla: o.id for o in db.scalars(select(OrganizacaoMilitar))}
    for sig in oms:
        if sig not in om_id:
            o = OrganizacaoMilitar(nome=sig, sigla=sig)
            db.add(o)
            db.flush()
            om_id[sig] = o.id

    # militares
    mil_id: dict[tuple[str, str], int] = {}
    for chave, m in militares.items():
        obj = Militar(nome_guerra=m["nome"], nome_completo=m["nome"],
                      posto_graduacao_id=pg[m["posto"]], om_id=om_id[m["om"]])
        db.add(obj)
        db.flush()
        mil_id[chave] = obj.id

    # escalas + 1 posto cada
    esc_id: dict[str, int] = {}
    posto_id: dict[str, int] = {}
    for aba, (nome, tp, tv) in ESCALAS.items():
        e = Escala(nome=nome, tem_preta=tp, tem_vermelha=tv)
        db.add(e)
        db.flush()
        esc_id[aba] = e.id
        p = Posto(escala_id=e.id, ordem=1, rotulo=None)
        db.add(p)
        db.flush()
        posto_id[aba] = p.id

    # concorrência: todos os pares
    ids = sorted(esc_id.values())
    for i in range(len(ids)):
        for j in range(i + 1, len(ids)):
            db.add(EscalaConcorrente(escala_menor_id=ids[i], escala_maior_id=ids[j]))

    # participações (dedup por militar+escala) + estado da fila (serviço sintético)
    part_vista: set[tuple[int, int]] = set()
    serv_vista: set[tuple[int, date]] = set()   # (posto_id, dia) — respeita uq_servico_posto_dia
    n_part = n_serv = colisoes = 0
    for r in registros:
        m = mil_id[(r.nome.upper(), r.om, r.posto)]
        e = esc_id[r.escala]
        if (m, e) not in part_vista:
            db.add(Participacao(militar_id=m, escala_id=e, ativo=True))
            part_vista.add((m, e))
            n_part += 1
        _, tp, tv = ESCALAS[r.escala]
        for cor, dia in ((Cor.PRETA, r.preta if tp else None),
                         (Cor.VERMELHA, r.vermelha if tv else None)):
            if dia is None:
                continue
            if (posto_id[r.escala], dia) in serv_vista:
                colisoes += 1
                continue
            serv_vista.add((posto_id[r.escala], dia))
            ini = datetime.combine(dia, time(8, 0))
            db.add(Servico(escala_id=e, posto_id=posto_id[r.escala], militar_id=m,
                           dia=dia, cor=cor, inicio_dt=ini,
                           termino_dt=ini + timedelta(hours=24)))
            n_serv += 1
    db.commit()
    return {"oms": len(oms), "militares": len(mil_id), "escalas": len(esc_id),
            "participacoes": n_part, "servicos": n_serv, "colisoes_servico": colisoes}


def main() -> None:
    ap = argparse.ArgumentParser(description="Migra a planilha do brigada (carga inicial).")
    ap.add_argument("--commit", action="store_true", help="grava de fato (padrão: dry-run)")
    ap.add_argument("--arquivo", default=ARQUIVO)
    args = ap.parse_args()

    caminho = Path(args.arquivo)
    if not caminho.exists():
        sys.exit(f"arquivo não encontrado: {caminho}")

    registros, avisos_ler = ler_planilha(caminho)
    oms, militares, avisos_cons = consolidar(registros)
    resumir(registros, oms, militares, avisos_ler, avisos_cons)

    if not args.commit:
        print("\n(dry-run: use --commit para gravar)")
        return

    db = SessionLocal()
    try:
        res = gravar(db, registros, oms, militares)
        print("\nGRAVADO:", res)
    finally:
        db.close()


if __name__ == "__main__":
    main()
